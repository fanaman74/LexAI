from __future__ import annotations
import io
import openpyxl
from extractors.common import ExtractionResult


def extract(data: bytes) -> ExtractionResult:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    md_parts: list[str] = []
    text_parts: list[str] = []
    sheet_names: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_names.append(sheet_name)
        md_parts.append(f"## Sheet: {sheet_name}")
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(cells)

        if not rows:
            continue

        header = rows[0]
        md_parts.append("| " + " | ".join(header) + " |")
        md_parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            padded = row + [""] * (len(header) - len(row))
            md_parts.append("| " + " | ".join(padded[:len(header)]) + " |")

        for row in rows:
            text_parts.append(" | ".join(row))

    wb.close()
    return ExtractionResult(
        text="\n".join(text_parts),
        markdown="\n".join(md_parts),
        metadata={"sheet_names": sheet_names, "sheet_count": len(sheet_names)},
    )
